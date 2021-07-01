from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
#1줄씩 입력

#한줄씩 입력
ws.append(["번호", "영어", "수학"])
for i in range(1,11): #10개 데이터 입력
    ws.append([i, randint(0,100), randint(0,100)]) #randint 0~100점 사이 값 입력

col_B = ws["B"] #영어 column, B열만 가져오기
# print(col_B)

# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"] #B~C까지의 모든 컬럼 데이터를 가지고 오기
# for cols in col_range : 
#     for cell in cols:
#         print(cell.value)

row_title = ws[1] #1번째 row 만 가지고 오기
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] #2~6번까지 가져오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end = " ")
#     print() #줄바꿈

# from openpyxl.utils.cell import coordinate_from_string #각 셀 좌표 정보 가져오기

# row_range = ws[2:ws.max_row] #2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end = " ")
#         # print(cell.coordinate, end = " ") #각 셀정보 가져오기
#         xy = coordinate_from_string(cell.coordinate) #A/10, AZ/250
#         # print(xy,end=" ")
#         print(xy[0],end="") #A 문자부분
#         print(xy[1],end=" ") #1 숫자부분
#     print()

#전체 rows
# print(tuple(ws.rows))
# print(tuple(ws.columns))

# for row in tuple(ws.rows):
#     print(row[0].value)

# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows(): #잔체 row
#     print(row[1].value)

# for column in ws.iter_cols(): #잔체 column
#     print(column[1].value)

# for row in tuple(ws.rows):
#     print(row[0].value)

#2~11번째 줄까지, 2번 열부터 3번째 열까지
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3): #전체 row
    print(row[0].value, row[1].value) #수학, 

wb.save("sample.xlsx")
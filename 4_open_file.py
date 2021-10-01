from openpyxl import load_workbook #파일불러오기
wb = load_workbook("These2_Company Coverage_210928__15")
ws = wb.active #활성화된 sheet

#cell 데이터 불러오기

for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(row = x, column = y).value, end = " ")
    print()

# #cell 개수를 모를 때
# for x in range(1,max_row + 1): #max_row: 최대 행수
#     for y in range(1,max_column + 1): #max_column 최대 열수
#         print(ws.cell(row = x, column = y).value, end = " ")
#     print()
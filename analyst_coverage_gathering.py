# 시트 결합하기
# https://hogni.tistory.com/18

#여러엑셀파일 결합하기
#엑셀 자동화 with 파이썬: 지정한 범위에서 데이터 찾아서 가져오기

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
# import numpy as np
# import xlwings as xw

#파일 위치 확인
input_folder = 'C:/Users/Peter/Documents/Python/RPA-1'
raw_data_dir = Path(input_folder)

#test* 형태의 파일 입력
excel_files = raw_data_dir.glob('test*')

#빈 데이터프레임 파일 생성
total_df = pd.DataFrame()

for excel_file in excel_files:
    # print(excel_file)
    #파일 내 모든 시트 내 모든 데이터 읽기
    df_all= pd.read_excel(excel_file, sheet_name=None)
    #df_all 내 데이터를 concatted_df로 모두 결합하기
    concatted_df = pd.concat(df_all, ignore_index = True)

    #각 파일 별로 데이터 취합
    total_df = total_df.append(concatted_df, ignore_index = True)

#취합한 파일을 입력할 위치 및 파일 생성
folder = 'C:/Users/Peter/Documents/Python/RPA-1/'
merged_excel_file = folder + 'test_merge.xlsx'

#데이터프레임 내 데이터를 엑셀 파일로 입력
total_df.to_excel(merged_excel_file, sheet_name = 'coverage', index=False)

#생성한 파일 및 경로 출력
# print("생성파일:", merged_excel_file)

# wb = load_workbook("test1.xlsx")

# ws = wb["1"]

# # # for x in range(11485, 18426):
# # for x in range(5, 10):
# #     for y in range(2,5):
# #         print(ws.cell(row=x, column=y).value, end=" ")
# #     print()


# # #cell 개수를 모를 때
# # for x in range(1,ws.max_row + 1): #max_row: 최대 행수
# #     for y in range(1,ws.max_column + 1): #max_column 최대 열수
# #         print(ws.cell(row = x, column = y).value, end = " ")
# #     print()

# # for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3): #전체 row
# for row in ws.iter_rows(): #전체 row
#     print(row) #수학, 




# # wb = Workbook() # 새 워크북 생성
# # ws = wb.active #현재 활성화된 sheet 가져옴
# # ws.title = "Coverage" #Sheet의 이름을 변경
# # wb.save("th2_analyst_coverage.xlsx")
# # wb.close()


# # new_ws = wb["NewSheet"] #Dict 형태로 Sheet 접근이 가능함
# # print(wb.sheetnames) #모든 시트 이름 확인

# # #sheet 복사
# # new_ws["A1"] = "Test"
# # target = wb.copy_worksheet(new_ws)
# # target.title = "Copied Sheet"

# # ws["A1"] = 1 #1값을 입력함

# # print(ws["A1"])
# # print(ws["A1"].value) #A1셀의 값을 출력함

# # ws.cell(row = 1, column = 1).value #A1셀
# # print(ws.cell(row = 1, column = 1).value ) #ws["A1"].value

# # from random import *

# # #반목문으로 랜덤 숫자 채우기

# # index = 1
# # for x in range(1, 11):
# #     for y in range(1,11):
# #         # ws.cell(row = x, column = y, value = randint(0,100))
# #         ws.cell(row = x, column = y, value = index)
# #         index += 1


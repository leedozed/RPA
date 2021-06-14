from openpyxl import Workbook
wb = Workbook() # 새 워크북 생성
ws = wb.create_sheet() #새로운 시트 기본 이름으로 생성
ws.title = "Mysheet" #sheet 이름 변경
ws.sheet_properties.tabColor = "ff3399" #색상 변경

ws1 = wb.create_sheet("YourSheet") #주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) #2번째 인덱스에 Sheet 생성

new_ws = wb["NewSheet"] #Dict 형태로 Sheet 접근이 가능함

print(wb.sheetnames) #모든 시트 이름 확인

#sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")

